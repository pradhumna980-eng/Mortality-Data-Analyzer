import os
from openpyxl import load_workbook, Workbook
from collections import defaultdict
import matplotlib.pyplot as plt

AGE_BANDS = [(18,25),(26,35),(36,45),(46,55),(56,65),(66,75),(76,120)]
MONTH_ORDER = ["Jan","Feb","Mar","Apr","May","Jun",
               "Jul","Aug","Sep","Oct","Nov","Dec"]

def get_age_band(age):
    for lo, hi in AGE_BANDS:
        if lo <= age <= hi:
            return f"{lo}-{hi}"
    return "Unknown"


# -------------------- PATIENT RECORD --------------------

class PatientRecord:
    def __init__(self, patient_id, age, disease, infection_month, death, years_until_death):
        self.patient_id = patient_id

        try:
            self.age = int(age)
        except:
            self.age = 0

        self.age_band = get_age_band(self.age)
        self.disease = str(disease) if disease else ""
        self.infection_month = infection_month if infection_month else "Unknown"

        try:
            self.death = int(death)
        except:
            self.death = 0

        try:
            self.years_until_death = int(years_until_death)
        except:
            self.years_until_death = None


# -------------------- DATASET --------------------

class MortalityDataset:
    def __init__(self, excel_file):
        self.records = []
        self.load_data(excel_file)

    def load_data(self, excel_file):
        if not os.path.exists(excel_file):
            print("Excel file not found.")
            return

        wb = load_workbook(excel_file)
        ws = wb.active

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            if row is None or len(row) < 6:
                continue
            self.records.append(PatientRecord(*row[:6]))

    def filter_band_and_disease(self, age_band, disease):
        return [
            r for r in self.records
            if r.age_band == age_band and r.disease.lower() == disease.lower()
        ]


# -------------------- ANALYZER --------------------

class ActuarialAnalyzer:
    def __init__(self, dataset):
        self.dataset = dataset

    def conditional_death_prob_3yr(self, age_band, disease):
        subset = self.dataset.filter_band_and_disease(age_band, disease)
        n = len(subset)

        if n == 0:
            return {}, 0

        year_counts = defaultdict(int)
        for r in subset:
            if r.death == 1 and r.years_until_death in [1, 2, 3]:
                year_counts[r.years_until_death] += 1

        # Correct conditional probability:
        # P(die in year Y | survived to start of year Y)
        probs = {}
        survivors = n
        for y in [1, 2, 3]:
            probs[y] = year_counts[y] / survivors if survivors > 0 else 0
            survivors -= year_counts[y]

        return probs, n

    def seasonal_mortality(self):
        exposure = defaultdict(int)
        deaths = defaultdict(int)

        for r in self.dataset.records:
            exposure[r.infection_month] += 1
            if r.death == 1:
                deaths[r.infection_month] += 1

        rates = {
            m: (deaths[m] / exposure[m] if exposure[m] > 0 else 0)
            for m in exposure
        }

        return exposure, deaths, rates

    def life_table(self):
        bands = sorted(
            set(r.age_band for r in self.dataset.records if r.age_band != "Unknown"),
            key=lambda x: int(x.split('-')[0])
        )

        l = 100000
        table = []

        for b in bands:
            band_records = [r for r in self.dataset.records if r.age_band == b]
            exposure = len(band_records)

            if exposure == 0:
                continue

            deaths = sum(r.death for r in band_records)
            qx = deaths / exposure
            px = 1 - qx
            dx = int(l * qx)
            Lx = l - dx / 2  # person-years lived in interval

            table.append({
                "AgeBand": b,
                "Exposure": exposure,
                "lx": l,
                "dx": dx,
                "qx": round(qx, 6),
                "px": round(px, 6),
                "Lx": round(Lx, 2),
            })

            l -= dx

        # Compute Tx and ex backwards
        for i in range(len(table) - 1, -1, -1):
            if i == len(table) - 1:
                table[i]["Tx"] = table[i]["Lx"]
            else:
                table[i]["Tx"] = table[i]["Lx"] + table[i + 1]["Tx"]
            table[i]["ex"] = round(table[i]["Tx"] / table[i]["lx"], 4) if table[i]["lx"] > 0 else 0

        return table

    def export_results(self, age_band, disease, outfile="mortality_results.xlsx"):
        probs, n = self.conditional_death_prob_3yr(age_band, disease)
        exposure, deaths, rates = self.seasonal_mortality()
        lifetable = self.life_table()

        # ---- EXCEL ----
        wb = Workbook()

        ws1 = wb.active
        ws1.title = "Conditional_Probabilities"
        ws1.append(["AgeBand", "Disease", "N", "Prob_Year1", "Prob_Year2", "Prob_Year3"])
        ws1.append([
            age_band, disease, n,
            probs.get(1, 0), probs.get(2, 0), probs.get(3, 0)
        ])

        ws2 = wb.create_sheet("Seasonal_Mortality")
        ws2.append(["Month", "Exposure", "Deaths", "MortalityRate"])
        for m in [m for m in MONTH_ORDER if m in exposure]:
            ws2.append([m, exposure[m], deaths[m], rates[m]])

        ws3 = wb.create_sheet("Life_Table")
        ws3.append(["AgeBand", "Exposure", "lx", "dx", "qx", "px", "Lx", "Tx", "ex"])
        for row in lifetable:
            ws3.append([
                row["AgeBand"], row["Exposure"],
                row["lx"], row["dx"], row["qx"], row["px"],
                row["Lx"], row["Tx"], row["ex"]
            ])

        wb.save(outfile)

        # ---- GRAPHS ----
        if n > 0:
            plt.figure()
            plt.bar([1, 2, 3], [probs.get(y, 0) for y in [1, 2, 3]])
            plt.xlabel("Years Until Death")
            plt.ylabel("Probability")
            plt.title(f"3-Year Conditional Death Probability\n{age_band} - {disease}")
            plt.savefig("conditional_probability.png")
            plt.close()

        if rates:
            plt.figure()
            months = [m for m in MONTH_ORDER if m in rates]
            plt.plot(months, [rates[m] for m in months])
            plt.xlabel("Infection Month")
            plt.ylabel("Mortality Rate")
            plt.title("Seasonal Mortality Rate")
            plt.savefig("seasonal_mortality.png")
            plt.close()

        if lifetable:
            plt.figure()
            plt.plot(
                [r["AgeBand"] for r in lifetable],
                [r["qx"] for r in lifetable]
            )
            plt.xlabel("Age Band")
            plt.ylabel("qx")
            plt.title("Life Table Mortality Rate by Age Band")
            plt.xticks(rotation=45)
            plt.tight_layout()
            plt.savefig("life_table_qx.png")
            plt.close()

        return outfile


# -------------------- MAIN --------------------

def main():
    # Portable path — works on any machine
    file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "patient_mortality_dataset.xlsx")

    data = MortalityDataset(file)

    if not data.records:
        print("No data loaded. Check that patient_mortality_dataset.xlsx is in the same folder.")
        return

    analyzer = ActuarialAnalyzer(data)

    try:
        age = int(input("Enter age: "))
    except:
        print("Invalid age. Please enter a number.")
        return

    disease = input("Enter disease: ")

    age_band = get_age_band(age)
    probs, n = analyzer.conditional_death_prob_3yr(age_band, disease)

    if n == 0:
        print(f"No matching records found for age band '{age_band}' and disease '{disease}'.")
        return

    print(f"\nAge band used: {age_band} (N={n})")
    print(f"P(death in year 1 | alive at start): {probs[1]:.4f}")
    print(f"P(death in year 2 | survived year 1): {probs[2]:.4f}")
    print(f"P(death in year 3 | survived year 2): {probs[3]:.4f}")

    print("\nLife Table:")
    print(f"{'AgeBand':<10} {'lx':>8} {'dx':>8} {'qx':>10} {'px':>10} {'Lx':>10} {'Tx':>12} {'ex':>8}")
    print("-" * 80)
    for row in analyzer.life_table():
        print(f"{row['AgeBand']:<10} {row['lx']:>8} {row['dx']:>8} {row['qx']:>10.4f} {row['px']:>10.4f} {row['Lx']:>10.1f} {row['Tx']:>12.1f} {row['ex']:>8.4f}")

    outfile = analyzer.export_results(age_band, disease)
    print(f"\nExcel exported to: {outfile}")
    print("Charts saved: conditional_probability.png, seasonal_mortality.png, life_table_qx.png")


if __name__ == "__main__":
    main()